import asyncio
import os
import re
from collections import defaultdict
from pathlib import Path
from urllib.parse import urljoin

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.async_api import Page, async_playwright


LIST_URL = "https://pk.mpei.ru/inform/list"
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = SCRIPT_DIR / "СФ_МЭИ_Сводка_для_руководства.xlsx"
HEADLESS = True


def clean_text(text: str) -> str:
    return " ".join((text or "").split())


def normalize_text(text: str) -> str:
    return (
        clean_text(text)
        .lower()
        .replace("ё", "е")
        .replace("–", "-")
        .replace("—", "-")
    )


def direction_key(direction: str) -> str:
    """Сопоставляет одинаковое направление между двумя страницами сайта."""
    text = clean_text(direction)
    code_match = re.search(r"\b\d{2}\.\d{2}\.\d{2}\b", text)

    if code_match:
        return code_match.group(0)

    return re.sub(r"[^а-яa-z0-9]+", "", normalize_text(text))


def competition_type(link_text: str) -> str | None:
    text = normalize_text(link_text)

    if (
        "основные места в рамках кцп" in text
        and "7 августа" in text
        and "бви" not in text
    ):
        return "Основной конкурс"

    return None


async def collect_main_links(page: Page) -> list[dict]:
    """Собирает только основной конкурс: бакалавриат, очная форма, Смоленск."""
    await page.goto(LIST_URL, wait_until="domcontentloaded", timeout=60_000)
    await page.wait_for_timeout(1200)

    collected = await page.evaluate(
        """
        () => {
            const result = [];
            let correctSectionReached = false;
            let insideSmolensk = false;

            const allElements = Array.from(document.querySelectorAll("body *"));

            for (const element of allElements) {
                const ownText = Array.from(element.childNodes)
                    .filter(node => node.nodeType === Node.TEXT_NODE)
                    .map(node => (node.textContent || "").trim())
                    .filter(Boolean)
                    .join(" ");

                const fullText = (element.innerText || "").trim();

                if (
                    ownText === "Бакалавриат очная форма обучения" ||
                    fullText === "Бакалавриат очная форма обучения"
                ) {
                    correctSectionReached = true;
                    insideSmolensk = false;
                    continue;
                }

                if (!correctSectionReached) continue;

                if (
                    ownText === "Бакалавриат очно-заочная форма обучения" ||
                    ownText === "Бакалавриат заочная форма обучения" ||
                    ownText === "Магистратура очная форма обучения" ||
                    fullText === "Бакалавриат очно-заочная форма обучения"
                ) {
                    break;
                }

                if (
                    ownText.includes('Смоленский филиал НИУ "МЭИ"') ||
                    fullText === 'Смоленский филиал НИУ "МЭИ" (г.Смоленск)'
                ) {
                    insideSmolensk = true;
                    continue;
                }

                if (
                    insideSmolensk &&
                    (
                        ownText.includes('Волжский филиал НИУ "МЭИ"') ||
                        ownText.includes('Душанбинский филиал НИУ "МЭИ"') ||
                        fullText.includes('Волжский филиал НИУ "МЭИ"')
                    )
                ) {
                    break;
                }

                if (!insideSmolensk || element.tagName !== "A") continue;

                const row = element.closest("tr");
                if (!row) continue;

                const cells = Array.from(row.querySelectorAll(":scope > td"));
                if (cells.length < 2) continue;

                result.push({
                    href: element.href,
                    linkText: element.innerText.trim(),
                    direction: cells[0].innerText.trim()
                });
            }

            return result;
        }
        """
    )

    links = []
    seen_urls = set()

    for item in collected:
        if competition_type(item["linkText"]) != "Основной конкурс":
            continue

        url = urljoin(LIST_URL, item["href"])
        if url in seen_urls:
            continue

        seen_urls.add(url)
        links.append(
            {
                "Направление": clean_text(item["direction"])
                or "Неизвестное направление",
                "Ссылка": url,
            }
        )

    return links


async def read_main_list(page: Page, url: str) -> dict:
    """
    Основной конкурс:
    - количество и минимальный балл: жирные зелёные;
    - кандидаты переполнения: жирные жёлтые с приоритетом.
    """
    await page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    await page.wait_for_timeout(700)

    return await page.evaluate(
        """
        () => {
            const GREEN = "rgb(204, 255, 204)";
            const YELLOW = "rgb(255, 255, 204)";

            function norm(text) {
                return (text || "")
                    .toLowerCase()
                    .replace(/ё/g, "е")
                    .replace(/\\s+/g, " ")
                    .trim();
            }

            function normalizedColor(element) {
                return getComputedStyle(element)
                    .backgroundColor
                    .replace(/\\s+/g, " ")
                    .trim();
            }

            function rowColor(row) {
                const own = normalizedColor(row);

                if (
                    own !== "rgba(0, 0, 0, 0)" &&
                    own !== "transparent"
                ) {
                    return own;
                }

                for (const cell of row.querySelectorAll(":scope > td")) {
                    const color = normalizedColor(cell);

                    if (
                        color !== "rgba(0, 0, 0, 0)" &&
                        color !== "transparent"
                    ) {
                        return color;
                    }
                }

                return null;
            }

            function isCodeBold(codeCell) {
                const weight = getComputedStyle(codeCell).fontWeight;

                if (weight === "bold" || weight === "bolder") {
                    return true;
                }

                const numeric = Number(weight);
                return Number.isFinite(numeric) && numeric >= 600;
            }

            function applicantFromRow(row) {
                const cells = Array.from(row.querySelectorAll(":scope > td"));

                for (const cell of cells) {
                    const text = cell.innerText.trim();

                    if (/^\\d{5,14}$/.test(text)) {
                        return {
                            code: text,
                            codeCell: cell,
                            cells
                        };
                    }
                }

                return null;
            }

            function findLogicalColumn(table, predicate) {
                const rows = Array.from(table.querySelectorAll("tr"));

                for (const row of rows.slice(0, 12)) {
                    const cells = Array.from(
                        row.querySelectorAll(":scope > th, :scope > td")
                    );

                    let logicalIndex = 0;

                    for (const cell of cells) {
                        if (predicate(norm(cell.innerText))) {
                            return logicalIndex;
                        }

                        logicalIndex += Math.max(
                            Number(cell.colSpan) || 1,
                            1
                        );
                    }
                }

                return -1;
            }

            function cellAtLogicalColumn(cells, logicalColumn) {
                if (logicalColumn < 0) return null;

                let logicalIndex = 0;

                for (const cell of cells) {
                    const span = Math.max(
                        Number(cell.colSpan) || 1,
                        1
                    );

                    if (
                        logicalColumn >= logicalIndex &&
                        logicalColumn < logicalIndex + span
                    ) {
                        return cell;
                    }

                    logicalIndex += span;
                }

                return null;
            }

            function parseNumber(text, minValue, maxValue) {
                const cleaned = (text || "")
                    .replace(/\\s+/g, "")
                    .replace(",", ".");

                if (!/^\\d+(\\.\\d+)?$/.test(cleaned)) return null;

                const value = Number(cleaned);

                return (
                    Number.isFinite(value) &&
                    value >= minValue &&
                    value <= maxValue
                )
                    ? value
                    : null;
            }

            function scoreFromRow(cells, column) {
                const scoreCell = cellAtLogicalColumn(cells, column);

                if (scoreCell) {
                    const value = parseNumber(
                        scoreCell.innerText,
                        0,
                        500
                    );

                    if (value !== null) return value;
                }

                const candidates = [];

                for (const cell of cells) {
                    const text = cell.innerText.trim();

                    if (/^\\d{5,14}$/.test(text)) continue;

                    const value = parseNumber(text, 100, 500);
                    if (value !== null) candidates.push(value);
                }

                return candidates.length
                    ? Math.max(...candidates)
                    : null;
            }

            function priorityFromRow(cells, column) {
                const priorityCell = cellAtLogicalColumn(cells, column);
                if (!priorityCell) return null;

                const value = parseNumber(
                    priorityCell.innerText,
                    1,
                    100
                );

                return value === null ? null : Math.trunc(value);
            }

            const accepted = [];
            const overflow = [];
            const seenGreen = new Set();
            const seenYellow = new Set();
            let globalRank = 0;

            for (const table of document.querySelectorAll("table")) {
                const scoreColumn = findLogicalColumn(
                    table,
                    text =>
                        text.includes("сумма конкурсных баллов") ||
                        text.includes("сумма баллов") ||
                        text.includes("конкурсный балл") ||
                        text === "баллы" ||
                        text === "балл"
                );

                let priorityColumn = findLogicalColumn(
                    table,
                    text =>
                        text.includes("приоритет зачисления") ||
                        text.includes("номер приоритета") ||
                        text.includes("№ приоритета")
                );

                if (priorityColumn < 0) {
                    priorityColumn = findLogicalColumn(
                        table,
                        text => text.includes("приоритет")
                    );
                }

                for (const row of table.querySelectorAll("tr")) {
                    const applicant = applicantFromRow(row);
                    if (!applicant) continue;

                    globalRank++;

                    const color = rowColor(row);
                    const bold = isCodeBold(applicant.codeCell);
                    if (!bold) continue;

                    const item = {
                        code: applicant.code,
                        score: scoreFromRow(
                            applicant.cells,
                            scoreColumn
                        ),
                        priority: priorityFromRow(
                            applicant.cells,
                            priorityColumn
                        ),
                        rank: globalRank
                    };

                    if (
                        color === GREEN &&
                        !seenGreen.has(item.code)
                    ) {
                        seenGreen.add(item.code);
                        accepted.push(item);
                    }

                    if (
                        color === YELLOW &&
                        !seenYellow.has(item.code)
                    ) {
                        seenYellow.add(item.code);
                        overflow.push(item);
                    }
                }
            }

            return {accepted, overflow};
        }
        """
    )


def build_summary(main_records: list[dict]) -> pd.DataFrame:
    main_by_key = {}
    overflow_occurrences = defaultdict(list)

    for record in main_records:
        direction = record["Направление"]
        key = direction_key(direction)
        accepted = record["accepted"]

        main_scores = [
            item["score"]
            for item in accepted
            if item.get("score") is not None
        ]

        main_by_key[key] = {
            "Направление": direction,
            "Основной конкурс": len(
                {item["code"] for item in accepted}
            ),
            "Минимальный балл основного конкурса": (
                min(main_scores) if main_scores else None
            ),
        }

        for item in record["overflow"]:
            overflow_occurrences[item["code"]].append(
                {
                    **item,
                    "direction": direction,
                    "direction_key": key,
                }
            )

    # Каждый жирный жёлтый УИП относится только к направлению
    # с наивысшим приоритетом.
    overflow_codes_by_key = defaultdict(set)

    for code, occurrences in overflow_occurrences.items():
        best = min(
            occurrences,
            key=lambda item: (
                item.get("priority")
                if item.get("priority") is not None
                else 10**9,
                item.get("rank", 10**9),
                item["direction"],
            ),
        )

        overflow_codes_by_key[best["direction_key"]].add(code)

    rows = []

    for key, main in main_by_key.items():
        rows.append(
            {
                "Направление": main["Направление"],
                "Основной конкурс": main["Основной конкурс"],
                "Переполнение": len(
                    overflow_codes_by_key.get(key, set())
                ),
                "Минимальный балл основного конкурса": main[
                    "Минимальный балл основного конкурса"
                ],
            }
        )

    summary = pd.DataFrame(
        rows,
        columns=[
            "Направление",
            "Основной конкурс",
            "Переполнение",
            "Минимальный балл основного конкурса",
        ],
    ).sort_values(
        "Направление",
        ignore_index=True,
    )

    total_row = pd.DataFrame(
        [
            {
                "Направление": "ИТОГО",
                "Основной конкурс": int(
                    summary["Основной конкурс"].sum()
                ),
                "Переполнение": int(
                    summary["Переполнение"].sum()
                ),
                "Минимальный балл основного конкурса": None,
            }
        ]
    )

    return pd.concat(
        [summary, total_row],
        ignore_index=True,
    )

def format_excel(output_path: Path) -> None:
    workbook = load_workbook(output_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )
    total_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in sheet.columns:
            max_length = max(
                len("" if cell.value is None else str(cell.value))
                for cell in column_cells
            )

            sheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(max(max_length + 2, 12), 70)

    if "Сводка" in workbook.sheetnames:
        sheet = workbook["Сводка"]

        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)
            cell.fill = total_fill

    workbook.save(output_path)


async def main() -> None:
    print("=" * 72)
    print("СФ МЭИ — СВОДКА ДЛЯ РУКОВОДСТВА")
    print("Основной конкурс и переполнение")
    print("=" * 72)

    main_records = []

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(
            headless=HEADLESS
        )
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            locale="ru-RU",
        )
        page = await context.new_page()

        try:
            print("1. Загружаю основной конкурс Смоленского филиала...")
            links = await collect_main_links(page)
            print(f"   Найдено направлений: {len(links)}")

            for index, item in enumerate(links, start=1):
                print(
                    f"   [{index}/{len(links)}] "
                    f"{item['Направление']}"
                )

                data = await read_main_list(
                    page,
                    item["Ссылка"],
                )

                main_records.append(
                    {
                        **item,
                        **data,
                    }
                )

        finally:
            await browser.close()

    if not main_records:
        print("Основной конкурс не получен.")
        return

    summary = build_summary(main_records)

    print()
    print("2. Формирую итоговую сводку...")

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:
        summary.to_excel(
            writer,
            sheet_name="Сводка",
            index=False,
        )

    format_excel(OUTPUT_FILE)

    print("=" * 72)
    print("ГОТОВО")
    print(OUTPUT_FILE)
    print("=" * 72)

    try:
        os.startfile(OUTPUT_FILE)
    except (AttributeError, OSError):
        pass


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nРабота остановлена пользователем.")
    except Exception as error:
        print("\nПроизошла ошибка:")
        print(error)
        print("\nСфотографируй окно или скопируй текст ошибки.")
    finally:
        input("\nНажми Enter, чтобы закрыть окно...")
